"""Tests for student CRUD, stage management, assignment, and import."""

from io import BytesIO

import pytest
from openpyxl import Workbook
from sqlalchemy import select

from app.models import OperationLog, Student
from app.routers.students import _is_within_dial_window


def test_dial_window_includes_end_minute_and_cross_midnight():
    assert _is_within_dial_window(23 * 60 + 59, "00:00", "23:59")
    assert _is_within_dial_window(23 * 60 + 30, "22:00", "01:00")
    assert _is_within_dial_window(30, "22:00", "01:00")
    assert not _is_within_dial_window(2 * 60, "22:00", "01:00")


@pytest.mark.asyncio
class TestCreateStudent:
    async def test_create_success(self, client, admin_headers):
        resp = await client.post(
            "/api/students",
            json={
                "name": "李四",
                "region": "湖里区",
                "guardian_phone": "13800138000",
            },
            headers=admin_headers,
        )
        assert resp.status_code == 200
        body = resp.json()
        assert body["code"] == 0
        assert body["data"]["name"] == "李四"

    async def test_create_no_region(self, client, admin_headers):
        resp = await client.post(
            "/api/students",
            json={
                "name": "王五",
                "guardian_phone": "13800138001",
            },
            headers=admin_headers,
        )
        assert resp.status_code == 200

    async def test_create_missing_name(self, client, admin_headers):
        resp = await client.post(
            "/api/students",
            json={"region": "测试"},
            headers=admin_headers,
        )
        assert resp.status_code == 422

    async def test_create_without_phone(self, client, admin_headers):
        resp = await client.post("/api/students", json={"name": "测试"}, headers=admin_headers)
        assert resp.status_code == 200
        body = resp.json()
        assert body["code"] == 1
        assert body["msg"] == "至少需要一个可拨电话"

    async def test_create_empty_body(self, client, admin_headers):
        resp = await client.post("/api/students", json={}, headers=admin_headers)
        assert resp.status_code == 422

    async def test_create_unauthorized(self, client):
        resp = await client.post("/api/students", json={"name": "赵六"})
        assert resp.status_code == 401

    async def test_create_unicode_name(self, client, admin_headers):
        resp = await client.post(
            "/api/students",
            json={
                "name": "𩷶测试",
                "region": "𩸽区",
                "guardian_phone": "13800138002",
            },
            headers=admin_headers,
        )
        assert resp.status_code == 200

    async def test_create_same_name_allowed_with_different_phone(self, client, admin_headers):
        """Students no longer require phone-based uniqueness."""
        resp1 = await client.post(
            "/api/students",
            json={
                "name": "重复电话",
                "guardian_phone": "13800138003",
            },
            headers=admin_headers,
        )
        assert resp1.json()["code"] == 0

        resp2 = await client.post(
            "/api/students",
            json={
                "name": "重复电话2",
                "guardian_phone": "13800138004",
            },
            headers=admin_headers,
        )
        assert resp2.json()["code"] == 0


@pytest.mark.asyncio
class TestListStudents:
    async def test_list_empty(self, client, admin_headers):
        resp = await client.get("/api/students", headers=admin_headers)
        body = resp.json()
        assert body["code"] == 0
        assert body["data"]["total"] == 0
        assert body["data"]["list"] == []

    async def test_list_with_data(self, client, admin_headers, sample_student):
        resp = await client.get("/api/students", headers=admin_headers)
        body = resp.json()
        assert body["data"]["total"] == 1
        assert body["data"]["list"][0]["name"] == "张三"

    async def test_list_pagination(self, client, admin_headers, db):
        from app.models import Student

        for i in range(15):
            db.add(Student(name=f"学生{i}", region="测试"))
        await db.commit()

        resp = await client.get("/api/students?page=1&page_size=10", headers=admin_headers)
        body = resp.json()
        assert len(body["data"]["list"]) == 10
        assert body["data"]["total"] == 15

        resp = await client.get("/api/students?page=2&page_size=10", headers=admin_headers)
        assert len(resp.json()["data"]["list"]) == 5

    async def test_list_filter_by_region(self, client, admin_headers, db):
        from app.models import Student

        db.add(Student(name="A", region="思明区"))
        db.add(Student(name="B", region="湖里区"))
        db.add(Student(name="C", region="思明区"))
        await db.commit()

        resp = await client.get("/api/students?region=思明区", headers=admin_headers)
        assert resp.json()["data"]["total"] == 2

    async def test_list_search_includes_guardian_phone_fields(self, client, admin_headers, db):
        from app.models import Student

        db.add(
            Student(
                name="第一监护人电话",
                guardian_phone="13960118706",
                guardian2_phone="",
            )
        )
        db.add(
            Student(
                name="第二监护人电话",
                guardian_phone="",
                guardian2_phone="18960100618",
            )
        )
        await db.commit()

        resp = await client.get("/api/students?q=18960100618", headers=admin_headers)
        body = resp.json()

        assert body["code"] == 0
        assert body["data"]["total"] == 1
        assert body["data"]["list"][0]["name"] == "第二监护人电话"

    async def test_list_search_normalizes_numeric_phone_query(self, client, admin_headers):
        resp = await client.post(
            "/api/students",
            json={
                "name": "格式化号码",
                "guardian2_phone": "189 6010-0618",
            },
            headers=admin_headers,
        )
        assert resp.json()["code"] == 0

        resp = await client.get("/api/students?q=18960100618", headers=admin_headers)
        body = resp.json()

        assert body["code"] == 0
        assert body["data"]["total"] == 1
        assert body["data"]["list"][0]["name"] == "格式化号码"
        assert body["data"]["list"][0]["guardian2_phone"] == "18960100618"
        assert body["data"]["list"][0]["guardian2_phone_raw"] is None

    async def test_list_search_hides_invalid_students_by_default(self, client, admin_headers, db):
        from app.models import Student, StudentStatus

        db.add(
            Student(
                name="黄妙娟",
                status=StudentStatus.child_not_want_study,
                status_detail="孩子不想读",
            )
        )
        await db.commit()

        resp = await client.get("/api/students?q=黄妙娟", headers=admin_headers)
        body = resp.json()

        assert body["code"] == 0
        assert body["data"]["total"] == 0

    async def test_list_status_filter_can_search_invalid_students(self, client, admin_headers, db):
        from app.models import Student, StudentStatus

        db.add(
            Student(
                name="未删除无效学生",
                status=StudentStatus.not_interested,
                status_detail="无意向",
            )
        )
        await db.commit()

        resp = await client.get("/api/students?status=无效&q=未删除无效学生", headers=admin_headers)
        body = resp.json()

        assert body["code"] == 0
        assert body["data"]["total"] == 1
        assert body["data"]["list"][0]["name"] == "未删除无效学生"
        assert body["data"]["list"][0]["status"] == "无效"
        assert body["data"]["list"][0]["status_detail"] == "无意向"

    async def test_list_search_excludes_invalid_student_after_delete(
        self, client, admin_headers, db
    ):
        from app.models import Student, StudentStatus

        student = Student(
            name="删除后搜不到",
            status=StudentStatus.not_interested,
            status_detail="无意向",
        )
        db.add(student)
        await db.commit()
        await db.refresh(student)

        delete_resp = await client.delete(f"/api/students/{student.id}", headers=admin_headers)
        assert delete_resp.json()["code"] == 0

        resp = await client.get("/api/students?q=删除后搜不到", headers=admin_headers)
        body = resp.json()

        assert body["code"] == 0
        assert body["data"]["total"] == 0

    async def test_list_returns_phone_fields_without_masking(self, client, admin_headers):
        resp = await client.post(
            "/api/students",
            json={
                "name": "脱敏号码",
                "guardian_phone": "13960118706",
                "guardian2_phone": "18960100618",
            },
            headers=admin_headers,
        )
        assert resp.json()["code"] == 0

        resp = await client.get("/api/students?q=13960118706", headers=admin_headers)
        item = resp.json()["data"]["list"][0]

        assert item["guardian_phone"] == "13960118706"
        assert item["guardian_phone_raw"] is None
        assert item["guardian2_phone"] == "18960100618"
        assert item["guardian2_phone_raw"] is None

    async def test_list_invalid_page(self, client, admin_headers):
        """Page has ge=1 validation, so -1 returns 422."""
        resp = await client.get("/api/students?page=-1", headers=admin_headers)
        assert resp.status_code == 422


@pytest.mark.asyncio
class TestImportStudents:
    async def test_import_requires_admin(self, client, agent_headers):
        wb = Workbook()
        ws = wb.active
        ws.append(["name", "phone"])
        ws.append(["Agent Import", "13900139000"])
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        resp = await client.post(
            "/api/students/import",
            headers=agent_headers,
            files={
                "file": (
                    "students.xlsx",
                    buf.getvalue(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

        assert resp.status_code == 403

    async def test_import_xlsx_skips_invalid_rows(self, client, admin_headers, db):
        from sqlalchemy import select

        from app.models import Student

        wb = Workbook()
        ws = wb.active
        ws.append(["name", "phone", "score", "guardian_name", "school_name"])
        ws.append(["Alice", "", "", "Parent A", "School A"])
        ws.append(["", "13900139000", "600", "Parent B", "School B"])
        ws.append(["Bob", "13800138000", "not-a-score", "Parent C", "School C"])
        ws.append(["Carol", "13700137000", 610, "", "School D"])
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        resp = await client.post(
            "/api/students/import",
            headers=admin_headers,
            files={
                "file": (
                    "students.xlsx",
                    buf.getvalue(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

        body = resp.json()
        assert body["code"] == 0
        assert body["data"]["imported"] == 1
        assert body["data"]["skipped"] == 3
        assert body["data"]["no_phone"] == 1
        assert body["data"]["no_phone_rows"] == [
            {"row": 2, "name": "Alice", "reason": "无电话数据"}
        ]

        result = await db.execute(select(Student).order_by(Student.name))
        students = result.scalars().all()
        assert [student.name for student in students] == ["Carol"]
        assert students[0].score == 610

    async def test_import_requires_xlsx(self, client, admin_headers):
        resp = await client.post(
            "/api/students/import",
            headers=admin_headers,
            files={"file": ("students.csv", b"name\nAlice\n", "text/csv")},
        )
        assert resp.json()["code"] == 1

    async def test_import_skips_unlabeled_row_without_phone(self, client, admin_headers, db):
        from sqlalchemy import select

        from app.models import Student

        wb = Workbook()
        ws = wb.active
        ws.append(["张三", "张父", "李母", "第一中学"])
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        resp = await client.post(
            "/api/students/import",
            headers=admin_headers,
            files={
                "file": (
                    "students.xlsx",
                    buf.getvalue(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

        body = resp.json()
        assert body["code"] == 0
        assert body["data"]["imported"] == 0
        assert body["data"]["no_phone"] == 1
        assert body["data"]["no_phone_rows"][0]["name"] == "张三"

        result = await db.execute(select(Student).where(Student.name == "张三"))
        assert result.scalar_one_or_none() is None


@pytest.mark.asyncio
class TestGetStudent:
    async def test_get_success(self, client, admin_headers, sample_student):
        resp = await client.get(f"/api/students/{sample_student.id}", headers=admin_headers)
        assert resp.json()["data"]["name"] == "张三"

    async def test_get_returns_phone_fields_without_masking(self, client, admin_headers, db):
        from app.models import Student

        student = Student(
            name="详情脱敏",
            guardian_phone="13960118706",
            guardian2_phone="18960100618",
        )
        db.add(student)
        await db.commit()
        await db.refresh(student)

        resp = await client.get(f"/api/students/{student.id}", headers=admin_headers)
        data = resp.json()["data"]

        assert data["guardian_phone"] == "13960118706"
        assert data["guardian_phone_raw"] is None
        assert data["guardian2_phone"] == "18960100618"
        assert data["guardian2_phone_raw"] is None

    async def test_reveal_phone_returns_plaintext_and_writes_audit_log(
        self, client, admin_headers, db
    ):
        from sqlalchemy import select

        from app.models import OperationLog, Student

        student = Student(
            name="审计取号",
            guardian_phone="13960118706",
            guardian2_phone="18960100618",
        )
        db.add(student)
        await db.commit()
        await db.refresh(student)

        resp = await client.get(f"/api/students/{student.id}/phone-plain", headers=admin_headers)
        data = resp.json()["data"]

        assert data == {
            "guardian_phone": "13960118706",
            "guardian2_phone": "18960100618",
        }
        logs = await db.execute(
            select(OperationLog).where(
                OperationLog.target_student_id == student.id,
                OperationLog.action == "查看明文电话",
            )
        )
        assert logs.scalar_one_or_none() is not None

    async def test_reveal_phone_requires_student_phone_operation_permission(
        self, client, db, normal_admin_user, normal_admin_headers
    ):
        student = Student(
            name="普通管理员取号",
            guardian_phone="13960118706",
            guardian2_phone="18960100618",
        )
        normal_admin_user.page_permissions = "leads_manage"
        db.add(student)
        await db.commit()

        resp = await client.get(
            f"/api/students/{student.id}/phone-plain", headers=normal_admin_headers
        )
        assert resp.status_code == 403

        normal_admin_user.operation_permissions = "student_phone"
        await db.commit()

        resp = await client.get(
            f"/api/students/{student.id}/phone-plain", headers=normal_admin_headers
        )
        assert resp.status_code == 200
        assert resp.json()["data"]["guardian_phone"] == "13960118706"

    async def test_get_not_found(self, client, admin_headers):
        """Backend raises HTTPException 404, not Response wrapper."""
        resp = await client.get("/api/students/99999", headers=admin_headers)
        assert resp.status_code == 404
        assert resp.json()["detail"] == "学生不存在"

    async def test_get_invalid_id(self, client, admin_headers):
        resp = await client.get("/api/students/abc", headers=admin_headers)
        assert resp.status_code in (404, 422)


@pytest.mark.asyncio
class TestUpdateStudent:
    async def test_update_rejects_clearing_all_phone_fields(self, client, admin_headers, db):
        from app.models import Student

        student = Student(
            name="禁止清空电话",
            guardian_phone="13800138005",
            guardian2_phone="13900139005",
        )
        db.add(student)
        await db.commit()
        await db.refresh(student)

        resp = await client.put(
            f"/api/students/{student.id}",
            json={"guardian_phone": "", "guardian2_phone": ""},
            headers=admin_headers,
        )
        body = resp.json()

        assert resp.status_code == 200
        assert body["code"] == 1
        assert body["msg"] == "至少需要一个可拨电话"
        await db.refresh(student)
        assert student.guardian_phone == "13800138005"
        assert student.guardian2_phone == "13900139005"

    async def test_update_status(self, client, admin_headers, sample_student):
        resp = await client.put(
            f"/api/students/{sample_student.id}",
            json={
                "status": "已联系",
            },
            headers=admin_headers,
        )
        assert resp.json()["code"] == 0

    async def test_update_new_lead_button_returns_default_not_contacted_status(
        self, client, admin_headers, sample_student
    ):
        resp = await client.put(
            f"/api/students/{sample_student.id}",
            json={"status": "新线索"},
            headers=admin_headers,
        )
        body = resp.json()

        assert body["code"] == 0
        assert body["data"]["status"] == "未联系"
        assert body["data"]["status_detail"] == ""

    async def test_update_interest_result_keeps_canonical_status_detail(
        self, client, admin_headers, sample_student
    ):
        resp = await client.put(
            f"/api/students/{sample_student.id}",
            json={"status": "非常有意向"},
            headers=admin_headers,
        )
        body = resp.json()

        assert body["code"] == 0
        assert body["data"]["status"] == "已联系"
        assert body["data"]["status_detail"] == "非常有意向"

    async def test_update_invalid_reason_status_keeps_reason_detail(
        self, client, admin_headers, sample_student
    ):
        resp = await client.put(
            f"/api/students/{sample_student.id}",
            json={"status": "孩子不想读"},
            headers=admin_headers,
        )
        body = resp.json()

        assert body["code"] == 0
        assert body["data"]["status"] == "无效"
        assert body["data"]["status_detail"] == "孩子不想读"
        assert body["data"]["invalid_reason"] == "孩子不想读"

    async def test_agent_cannot_write_call_result_without_recent_dial_log(
        self, client, db, agent_user, agent_headers
    ):
        from app.models import Student, StudentStatus

        student = Student(
            name="未拨号改状态",
            assigned_to=agent_user.id,
            status=StudentStatus.not_contacted,
            guardian_phone="13800138000",
        )
        db.add(student)
        await db.commit()
        await db.refresh(student)

        resp = await client.put(
            f"/api/students/{student.id}",
            json={"status": "无意向"},
            headers=agent_headers,
        )

        assert resp.status_code == 403
        assert "请先通过系统拨号按钮拨打" in resp.json()["detail"]

    async def test_agent_can_write_call_result_after_recent_dial_log(
        self, client, db, agent_user, agent_headers
    ):
        from app.models import DialLog, Student, StudentStatus

        student = Student(
            name="已拨号改状态",
            assigned_to=agent_user.id,
            status=StudentStatus.not_contacted,
            guardian_phone="13800138001",
        )
        db.add(student)
        await db.flush()
        db.add(DialLog(student_id=student.id, agent_id=agent_user.id))
        await db.commit()
        await db.refresh(student)

        resp = await client.put(
            f"/api/students/{student.id}",
            json={"status": "无意向"},
            headers=agent_headers,
        )

        assert resp.status_code == 200
        body = resp.json()
        assert body["code"] == 0
        assert body["data"]["status"] == "无效"
        assert body["data"]["status_detail"] == "无意向"

    async def test_agent_cannot_write_call_result_after_stale_dial_log(
        self, client, db, agent_user, agent_headers
    ):
        from datetime import timedelta

        from app.models import DialLog, Student, StudentStatus
        from app.utils import utcnow

        student = Student(
            name="过期拨号改状态",
            assigned_to=agent_user.id,
            status=StudentStatus.not_contacted,
            guardian_phone="13800138002",
        )
        db.add(student)
        await db.flush()
        db.add(
            DialLog(
                student_id=student.id,
                agent_id=agent_user.id,
                dialed_at=utcnow() - timedelta(hours=25),
            )
        )
        await db.commit()
        await db.refresh(student)

        resp = await client.put(
            f"/api/students/{student.id}",
            json={"status": "无意向"},
            headers=agent_headers,
        )

        assert resp.status_code == 403
        assert "请先通过系统拨号按钮拨打" in resp.json()["detail"]

    async def test_agent_can_backfill_call_result_for_already_contacted_student_without_dial_log(
        self, client, db, agent_user, agent_headers
    ):
        from app.models import Student, StudentStatus

        student = Student(
            name="已联系补录结果",
            assigned_to=agent_user.id,
            status=StudentStatus.contacted,
            guardian_phone="13800138003",
        )
        db.add(student)
        await db.commit()
        await db.refresh(student)

        resp = await client.put(
            f"/api/students/{student.id}",
            json={"status": "无意向"},
            headers=agent_headers,
        )

        assert resp.status_code == 200
        body = resp.json()
        assert body["code"] == 0
        assert body["data"]["status"] == "无效"
        assert body["data"]["status_detail"] == "无意向"

    async def test_admin_can_write_call_result_without_dial_log(
        self, client, sample_student, admin_headers
    ):
        resp = await client.put(
            f"/api/students/{sample_student.id}",
            json={"status": "无意向"},
            headers=admin_headers,
        )

        assert resp.status_code == 200
        assert resp.json()["code"] == 0

    async def test_get_phone_deduplicates_immediate_repeat_dial_log(
        self, client, db, agent_user, agent_headers
    ):
        from sqlalchemy import select

        from app.models import DialLog, Student, StudentStatus, SystemConfig

        student = Student(
            name="重复拨号学生",
            assigned_to=agent_user.id,
            status=StudentStatus.not_contacted,
            guardian_phone="13800138000",
        )
        db.add(student)
        db.add(SystemConfig(key="dial_window_start", value="00:00"))
        db.add(SystemConfig(key="dial_window_end", value="23:59"))
        await db.commit()
        await db.refresh(student)

        first = await client.get(f"/api/students/phone/{student.id}", headers=agent_headers)
        second = await client.get(f"/api/students/phone/{student.id}", headers=agent_headers)

        assert first.status_code == 200
        assert second.status_code == 200
        assert first.json()["code"] == 0
        assert second.json()["code"] == 0
        rows = (
            (
                await db.execute(
                    select(DialLog).where(
                        DialLog.student_id == student.id,
                        DialLog.agent_id == agent_user.id,
                    )
                )
            )
            .scalars()
            .all()
        )
        assert len(rows) == 1

    async def test_manual_intent_writes_operation_log(
        self, client, admin_headers, sample_student, db
    ):
        from sqlalchemy import select

        from app.models import OperationLog

        resp = await client.put(
            f"/api/students/{sample_student.id}",
            json={"intent_level": "A"},
            headers=admin_headers,
        )
        assert resp.json()["code"] == 0
        r = await db.execute(
            select(OperationLog).where(
                OperationLog.target_student_id == sample_student.id,
                OperationLog.action == "手动评级",
            )
        )
        assert len(r.scalars().all()) >= 1

    async def test_update_invalid_intent(self, client, admin_headers, sample_student):
        resp = await client.put(
            f"/api/students/{sample_student.id}",
            json={"intent_level": "无效等级"},
            headers=admin_headers,
        )
        # Pydantic now validates at request boundary (422) instead of app-level code=1
        assert resp.status_code == 422 or resp.json().get("code") == 1
        assert "意向" in str(resp.json()) or "无效" in str(resp.json())

    async def test_update_stage(self, client, admin_headers, sample_student):
        """Historical bug: stage enum serialization."""
        resp = await client.put(
            f"/api/students/{sample_student.id}/stage",
            json={
                "stage": "有意向",
            },
            headers=admin_headers,
        )
        assert resp.json()["code"] == 0
        assert resp.json()["data"]["stage"] == "有意向"

    async def test_update_stage_accepts_home_visit_stage(
        self, client, admin_headers, sample_student
    ):
        resp = await client.put(
            f"/api/students/{sample_student.id}/stage",
            json={"stage": "待家访"},
            headers=admin_headers,
        )

        assert resp.status_code == 200
        assert resp.json()["code"] == 0
        assert resp.json()["data"]["stage"] == "待家访"

    async def test_list_filter_accepts_home_visit_stage(self, client, db, admin_headers):
        from app.models import Student, StudentStage

        db.add_all(
            [
                Student(name="待家访学生", stage=StudentStage.home_visit_pending),
                Student(name="普通阶段学生", stage=StudentStage.initial_contact),
            ]
        )
        await db.commit()

        resp = await client.get(
            "/api/students",
            params={"stage": "待家访"},
            headers=admin_headers,
        )

        assert resp.status_code == 200
        body = resp.json()
        assert body["code"] == 0
        assert [item["name"] for item in body["data"]["list"]] == ["待家访学生"]

    async def test_list_filter_merges_legacy_visit_stage(self, client, db, admin_headers):
        from app.models import Student, StudentStage

        db.add_all(
            [
                Student(name="旧预约参观学生", stage=StudentStage.visit_scheduled),
                Student(name="新到校预约学生", stage=StudentStage.campus_visit_scheduled),
            ]
        )
        await db.commit()

        resp = await client.get(
            "/api/students",
            params={"stage": "到校参观已安排"},
            headers=admin_headers,
        )

        assert resp.status_code == 200
        body = resp.json()
        assert body["code"] == 0
        rows = body["data"]["list"]
        assert {item["name"] for item in rows} == {"旧预约参观学生", "新到校预约学生"}
        assert {item["stage"] for item in rows} == {"到校参观已安排"}

    async def test_enrolled_student_cannot_be_reopened_by_generic_status_update(
        self, client, db, admin_headers
    ):
        from app.models import Student, StudentStage, StudentStatus

        student = Student(
            name="已报名保护",
            status=StudentStatus.enrolled,
            stage=StudentStage.enrolled,
        )
        db.add(student)
        await db.commit()
        await db.refresh(student)

        resp = await client.put(
            f"/api/students/{student.id}",
            json={"status": "未联系"},
            headers=admin_headers,
        )

        assert resp.status_code == 200
        assert resp.json()["code"] == 1
        assert "已报名" in resp.json()["msg"]
        await db.refresh(student)
        assert student.status == StudentStatus.enrolled
        assert student.stage == StudentStage.enrolled

    async def test_update_stage_to_enrolled_writes_operation_log(
        self, client, db, admin_headers, sample_student
    ):
        resp = await client.put(
            f"/api/students/{sample_student.id}/stage",
            json={"stage": "已报名"},
            headers=admin_headers,
        )

        assert resp.json()["code"] == 0
        assert resp.json()["data"]["stage"] == "已报名"
        assert resp.json()["data"]["status"] == "已报名"

        log = (
            await db.execute(
                select(OperationLog).where(
                    OperationLog.target_student_id == sample_student.id,
                    OperationLog.action == "修改状态",
                )
            )
        ).scalar_one()
        assert log.old_status == "未联系"
        assert log.new_status == "已报名"
        assert "阶段 初次联系 → 已报名" in log.content
        assert "状态 未联系 → 已报名" in log.content

    async def test_set_enroll_info_writes_operation_log(
        self, client, db, admin_headers, sample_student
    ):
        resp = await client.put(
            f"/api/students/{sample_student.id}/enroll",
            json={"program": "护理", "deposit": 500, "enrolled_at": "2026-06-30"},
            headers=admin_headers,
        )

        assert resp.json()["code"] == 0
        assert resp.json()["data"]["program"] == "护理"

        log = (
            await db.execute(
                select(OperationLog).where(
                    OperationLog.target_student_id == sample_student.id,
                    OperationLog.action == "报名登记",
                )
            )
        ).scalar_one()
        assert log.old_status == "未联系"
        assert log.new_status == "已报名"
        assert "状态 未联系 → 已报名" in log.content
        assert "阶段 初次联系 → 已报名" in log.content
        assert "报名日 - → 2026-06-30" in log.content
        assert "专业 - → 护理" in log.content
        assert "定金 - → 500.0" in log.content

    async def test_update_invalid_stage(self, client, admin_headers, sample_student):
        """Historical bug: invalid stage crashed with 500. Now returns code=1."""
        resp = await client.put(
            f"/api/students/{sample_student.id}/stage",
            json={
                "stage": "无效阶段",
            },
            headers=admin_headers,
        )
        assert resp.status_code == 200
        assert resp.json()["code"] == 1
        assert "无效" in resp.json()["msg"]

    async def test_update_not_found(self, client, admin_headers):
        resp = await client.put(
            "/api/students/99999",
            json={"name": "新名字"},
            headers=admin_headers,
        )
        assert resp.status_code == 404
        assert resp.json()["detail"] == "学生不存在"


@pytest.mark.asyncio
class TestAssignStudent:
    async def test_manual_assign_writes_operation_log(
        self, client, db, admin_headers, sample_student, agent_user
    ):
        resp = await client.post(
            "/api/students/assign",
            json={"student_ids": [sample_student.id], "agent_id": agent_user.id},
            headers=admin_headers,
        )

        assert resp.json()["code"] == 0
        await db.refresh(sample_student)
        assert sample_student.assigned_to == agent_user.id
        log = (
            await db.execute(
                select(OperationLog).where(
                    OperationLog.target_student_id == sample_student.id,
                    OperationLog.action == "手动分配",
                )
            )
        ).scalar_one()
        assert log.content == f"分配给话务员 {agent_user.id}"

    async def test_manual_assign_writes_batch_summary_log(
        self, client, db, admin_headers, agent_user
    ):
        from app.models import Student, StudentStatus

        students = [
            Student(name="批量学生1", status=StudentStatus.not_contacted),
            Student(name="批量学生2", status=StudentStatus.not_contacted),
        ]
        db.add_all(students)
        await db.commit()
        for student in students:
            await db.refresh(student)

        resp = await client.post(
            "/api/students/assign",
            json={"student_ids": [s.id for s in students], "agent_id": agent_user.id},
            headers=admin_headers,
        )

        assert resp.json()["code"] == 0
        summary = (
            await db.execute(
                select(OperationLog).where(
                    OperationLog.action == "批量分配",
                    OperationLog.target_student_id.is_(None),
                )
            )
        ).scalar_one()
        assert "共 2 名" in summary.content
        assert f"话务员 {agent_user.id}" in summary.content
        assert "批量学生1、批量学生2" in summary.content

    async def test_manual_assign_rejects_enrolled_students(
        self, client, db, admin_headers, agent_user
    ):
        from app.models import Student, StudentStage, StudentStatus

        student = Student(
            name="已报名不能分配",
            status=StudentStatus.enrolled,
            stage=StudentStage.enrolled,
        )
        db.add(student)
        await db.commit()
        await db.refresh(student)

        resp = await client.post(
            "/api/students/assign",
            json={"student_ids": [student.id], "agent_id": agent_user.id},
            headers=admin_headers,
        )

        assert resp.status_code == 200
        assert resp.json()["code"] == 1
        assert "已报名" in resp.json()["msg"]
        await db.refresh(student)
        assert student.assigned_to is None

    async def test_auto_assign_skips_unassigned_enrolled_students(
        self, client, db, admin_headers, agent_user
    ):
        from app.models import Student, StudentStage, StudentStatus

        enrolled = Student(
            name="未分配已报名",
            status=StudentStatus.enrolled,
            stage=StudentStage.enrolled,
        )
        active = Student(name="未分配有效", status=StudentStatus.not_contacted)
        db.add_all([enrolled, active])
        await db.commit()

        resp = await client.post("/api/students/auto-assign", headers=admin_headers)

        assert resp.status_code == 200
        assert resp.json()["code"] == 0
        assert resp.json()["data"]["total_assigned"] == 1
        await db.refresh(enrolled)
        await db.refresh(active)
        assert enrolled.assigned_to is None
        assert active.assigned_to == agent_user.id

    async def test_auto_assign_writes_batch_summary_log(
        self, client, db, admin_headers, agent_user
    ):
        from app.models import Student, StudentStatus

        students = [
            Student(name="自动学生1", status=StudentStatus.not_contacted),
            Student(name="自动学生2", status=StudentStatus.not_contacted),
        ]
        db.add_all(students)
        await db.commit()

        resp = await client.post("/api/students/auto-assign", headers=admin_headers)

        assert resp.json()["code"] == 0
        summary = (
            await db.execute(
                select(OperationLog).where(
                    OperationLog.action == "自动分配汇总",
                    OperationLog.target_student_id.is_(None),
                )
            )
        ).scalar_one()
        assert "共 2 名" in summary.content
        assert f"{agent_user.id}:2" in summary.content
        assert "自动学生1、自动学生2" in summary.content


@pytest.mark.asyncio
class TestDeleteStudent:
    async def test_delete_success(self, client, admin_headers, sample_student):
        resp = await client.delete(f"/api/students/{sample_student.id}", headers=admin_headers)
        assert resp.json()["code"] == 0

    async def test_delete_requires_admin(self, client, agent_headers, sample_student):
        resp = await client.delete(f"/api/students/{sample_student.id}", headers=agent_headers)
        assert resp.status_code == 403

    async def test_delete_requires_super_admin(self, client, normal_admin_headers, sample_student):
        resp = await client.delete(
            f"/api/students/{sample_student.id}", headers=normal_admin_headers
        )
        assert resp.status_code == 403

    async def test_delete_not_found(self, client, admin_headers):
        resp = await client.delete("/api/students/99999", headers=admin_headers)
        assert resp.status_code == 404
        assert resp.json()["detail"] == "学生不存在"
