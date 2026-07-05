import asyncio
import logging
import uuid
from datetime import date, datetime, timedelta, timezone
from io import BytesIO
from itertools import chain
from zipfile import BadZipFile

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from pydantic import BaseModel, Field, field_validator
from sqlalchemy import delete, func, insert, or_, select, update
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.ext.asyncio import AsyncSession

from app.auth import (
    ADMIN_OP_ENROLLMENT_SETTLEMENT,
    ADMIN_OP_STUDENT_ASSIGN,
    ADMIN_OP_STUDENT_CREATE,
    ADMIN_OP_STUDENT_DELETE,
    ADMIN_OP_STUDENT_EDIT,
    ADMIN_OP_STUDENT_IMPORT,
    ADMIN_OP_STUDENT_PHONE,
    ADMIN_PAGE_ENROLLMENT_SETTLEMENT,
    ADMIN_PAGE_LEADS_MANAGE,
    get_current_user,
    require_admin,
    require_agent,
    require_operation_permission,
    require_page_permission,
    user_has_operation_permission,
    user_has_page_permission,
)
from app.database import get_db
from app.dial_guard import require_recent_agent_dial
from app.models import (
    Call,
    CampusVisitTask,
    DialLog,
    EnrollmentRecord,
    EnrollmentSubStage,
    FollowUp,
    HomeVisitTask,
    IntentLevel,
    LeadViewLog,
    Note,
    OperationLog,
    Student,
    StudentStage,
    StudentStatus,
    SystemConfig,
    User,
    UserRole,
    Visit,
)
from app.permissions import (
    apply_student_scope,
    get_accessible_student,
    get_student_or_404,
    is_admin,
)
from app.pushplus import notify_a_level_change_background
from app.region_extractor import extract_region
from app.schemas import EnrollInfo, Response, StageUpdate, StudentCreate, StudentUpdate
from app.status_policy import (
    canonical_status_value,
    canonical_student_status,
    normalize_status_for_write,
    status_detail_for_write,
    status_detail_value,
    statuses_for_canonical,
)
from app.student_import import (
    MAX_STUDENT_IMPORT_BYTES,
    build_import_header_map,
    is_empty_import_row,
    parse_import_row,
)
from app.task_stats import ACTIVE_TASK_STATUSES, TERMINAL_STUDENT_STATUSES
from app.utils import (
    assignment_state_label,
    is_phone_query,
    make_assignment_rollback_note,
    make_batch_id,
    make_operation_log,
    mask_phone,
    normalize_phone,
    today_cst_as_utc,
    utcnow,
)

router = APIRouter(prefix="/api/students", tags=["学生"])

logger = logging.getLogger(__name__)


STAGE_ORDER = [
    "初次联系",
    "有意向",
    "已送资料",
    "待家访",
    "家访已安排",
    "家访完成",
    "待到校参观",
    "到校参观已安排",
    "已到校参观",
    "已报名",
]

def _require_admin_leads_manage(current_user: User) -> None:
    if is_admin(current_user) and not user_has_page_permission(
        current_user, ADMIN_PAGE_LEADS_MANAGE
    ):
        raise HTTPException(status_code=403, detail="无权访问该管理模块")


def _require_admin_operation(current_user: User, permission: str) -> None:
    if is_admin(current_user) and not user_has_operation_permission(current_user, permission):
        raise HTTPException(status_code=403, detail="无权执行该操作")


ADMIN_STUDENT_UPDATE_FIELDS = {
    "name",
    "status",
    "intent_level",
    "assigned_to",
    "join_reasons",
    "region",
    "stage",
    "enrolled_at",
    "program",
    "deposit",
    "score",
    "guardian_name",
    "guardian_phone",
    "guardian2_name",
    "guardian2_phone",
    "school_name",
    "school_address",
    "need_help",
}
AGENT_STUDENT_UPDATE_FIELDS = {
    "status",
    "intent_level",
    "join_reasons",
    "stage",
    "need_help",
    "score",
}
CALL_RESULT_STATUSES = {
    StudentStatus.contacted,
    StudentStatus.not_reached,
    StudentStatus.pending_visit,
    StudentStatus.invalid,
}
CALL_RESULT_STATUS_DETAILS = {
    "非常有意向",
    "意向了解加微",
    "高分段",
    "无意向",
    "孩子不想读",
    "空号",
    "其他",
}


def next_stage(current: str) -> str | None:
    try:
        idx = STAGE_ORDER.index(current)
        return STAGE_ORDER[idx + 1] if idx + 1 < len(STAGE_ORDER) else None
    except ValueError:
        return None


def _enum_or_error(enum_cls, value: str, label: str):
    try:
        return enum_cls(value)
    except ValueError:
        raise ValueError(f"无效的{label}: {value}")


def _has_any_phone(guardian_phone: str | None, guardian2_phone: str | None) -> bool:
    return bool((guardian_phone or "").strip() or (guardian2_phone or "").strip())


def _dedupe_contact_phones(
    guardian_phone: str | None, guardian2_phone: str | None
) -> tuple[str, str]:
    phone = normalize_phone(guardian_phone)
    phone2 = normalize_phone(guardian2_phone)
    if phone and phone2 and phone == phone2:
        phone2 = ""
    return phone, phone2


def _display_stage(stage: StudentStage) -> str:
    if stage == StudentStage.visit_scheduled:
        return StudentStage.campus_visit_scheduled.value
    if stage == StudentStage.visited:
        return StudentStage.campus_visit_arrived.value
    return stage.value


def _stage_filter_values(stage: StudentStage) -> list[StudentStage]:
    if stage == StudentStage.campus_visit_scheduled:
        return [StudentStage.campus_visit_scheduled, StudentStage.visit_scheduled]
    if stage == StudentStage.campus_visit_arrived:
        return [StudentStage.campus_visit_arrived, StudentStage.visited]
    return [stage]


def _student_payload(student: Student) -> dict:
    status = canonical_status_value(student.status)
    detail = status_detail_value(student.status, getattr(student, "status_detail", ""))
    intent_level = student.intent_level.value
    stage = _display_stage(student.stage)
    payload = {
        "id": student.id,
        "name": student.name,
        "region": student.region,
        "assigned_to": student.assigned_to,
        "status": status,
        "status_detail": detail,
        "invalid_reason": detail if status == StudentStatus.invalid.value else "",
        "intent_level": intent_level,
        "stage": stage,
        "join_reasons": student.join_reasons,
        "case_no": student.case_no,
        "need_help": student.need_help,
        "score": student.score,
        "guardian_name": student.guardian_name,
        "guardian_phone": mask_phone(student.guardian_phone),
        "guardian_phone_raw": None,
        "guardian2_name": student.guardian2_name,
        "guardian2_phone": mask_phone(student.guardian2_phone),
        "guardian2_phone_raw": None,
        "school_name": student.school_name,
        "school_address": student.school_address,
        "enrolled_at": str(student.enrolled_at) if student.enrolled_at else None,
        "program": student.program,
        "deposit": student.deposit,
        "expired_at": str(student.expired_at) if student.expired_at else None,
        "enrollment_substage": student.enrollment_substage.value
        if student.enrollment_substage
        else None,
        "assigned_at": str(student.assigned_at) if student.assigned_at else None,
        "created_at": str(student.created_at),
        "updated_at": str(student.updated_at),
    }
    return payload


def _is_call_result_write(status: StudentStatus, status_detail: str | None) -> bool:
    canonical_status = canonical_student_status(status)
    detail = (status_detail or "").strip()
    return canonical_status in CALL_RESULT_STATUSES or detail in CALL_RESULT_STATUS_DETAILS


def _allows_call_result_backfill_without_recent_dial(
    old_status: StudentStatus | str | None,
) -> bool:
    return canonical_student_status(old_status) == StudentStatus.contacted


def _is_enrolled_student(student: Student) -> bool:
    return (
        canonical_student_status(student.status) == StudentStatus.enrolled
        or student.stage == StudentStage.enrolled
    )


@router.post("/import")
async def import_students_excel(
    file: UploadFile = File(...),
    default_agent_id: int | None = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_page_permission(ADMIN_PAGE_LEADS_MANAGE)),
):
    _require_admin_operation(current_user, ADMIN_OP_STUDENT_IMPORT)
    filename = file.filename or ""
    if not filename.lower().endswith(".xlsx"):
        return Response.error(code=1, msg="仅支持 .xlsx 文件")

    workbook = None
    try:
        contents = await file.read()
        if len(contents) > MAX_STUDENT_IMPORT_BYTES:
            max_mb = MAX_STUDENT_IMPORT_BYTES // (1024 * 1024)
            return Response.error(code=1, msg=f"文件过大，请小于 {max_mb}MB")
        if not contents:
            return Response.error(code=1, msg="上传文件为空")

        if default_agent_id is not None:
            agent_result = await db.execute(
                select(User.id).where(User.id == default_agent_id, User.is_active)
            )
            if not agent_result.scalar_one_or_none():
                return Response.error(code=1, msg="默认话务员不存在或已禁用")

        workbook = load_workbook(filename=BytesIO(contents), read_only=True, data_only=True)
        ws = workbook.active
        rows = ws.iter_rows(values_only=True)
        header_row = next(rows, None)
        if not header_row:
            return Response.error(code=1, msg="Excel 文件没有表头")

        header_map = build_import_header_map(header_row)
        data_start_row = 2
        if "name" not in header_map:
            rows = chain([header_row], rows)
            header_map = {}
            data_start_row = 1

        parsed_rows = []  # (row_idx, parsed_dict)
        skipped_rows = []
        assigned_at = utcnow() if default_agent_id is not None else None

        for row_idx, row in enumerate(rows, start=data_start_row):
            if is_empty_import_row(row):
                continue

            parsed, error = parse_import_row(row, header_map)
            if error:
                skipped_rows.append({"row": row_idx, "reason": error})
                continue

            parsed_rows.append((row_idx, parsed))

        # Batch-check for duplicate phone numbers against existing DB records.
        # 同一行的 guardian_phone / guardian2_phone 都纳入查重；库里命中任一字段视为已存在。
        phones_to_check: set[str] = set()
        for _, p in parsed_rows:
            for key in ("guardian_phone", "guardian2_phone"):
                val = p.get(key) or ""
                if val:
                    phones_to_check.add(val)

        existing_phones: set[str] = set()
        if phones_to_check:
            existing_r = await db.execute(
                select(Student.guardian_phone, Student.guardian2_phone).where(
                    or_(
                        Student.guardian_phone.in_(phones_to_check),
                        Student.guardian2_phone.in_(phones_to_check),
                    )
                )
            )
            for g1, g2 in existing_r.all():
                if g1:
                    existing_phones.add(g1)
                if g2:
                    existing_phones.add(g2)

        seen_in_file: set[str] = set()
        student_rows = []
        no_phone_rows = []
        for row_idx, parsed in parsed_rows:
            phone = parsed.get("guardian_phone", "") or ""
            phone2 = parsed.get("guardian2_phone", "") or ""
            row_phones = [p for p in (phone, phone2) if p]

            if not row_phones:
                item = {
                    "row": row_idx,
                    "name": parsed["name"],
                    "reason": "无电话数据",
                }
                no_phone_rows.append(item)
                skipped_rows.append(item)
                continue

            dup_db = next((p for p in row_phones if p in existing_phones), None)
            if dup_db:
                skipped_rows.append(
                    {"row": row_idx, "reason": f"手机号已存在（库中已有该学员）: {dup_db}"}
                )
                continue
            dup_file = next((p for p in row_phones if p in seen_in_file), None)
            if dup_file:
                skipped_rows.append(
                    {"row": row_idx, "reason": f"手机号在本次导入中重复: {dup_file}"}
                )
                continue
            seen_in_file.update(row_phones)

            student_rows.append(
                {
                    "name": parsed["name"],
                    "region": parsed["region"],
                    "assigned_to": default_agent_id,
                    "assigned_at": assigned_at,
                    "score": parsed["score"],
                    "guardian_name": parsed["guardian_name"],
                    "guardian_phone": phone,
                    "guardian2_name": parsed["guardian2_name"],
                    "guardian2_phone": phone2,
                    "school_name": parsed["school_name"],
                    "school_address": parsed["school_address"],
                    "program": parsed["program"],
                    "join_reasons": parsed["join_reasons"],
                    "status": StudentStatus.not_contacted,
                    "status_detail": "",
                    "intent_level": IntentLevel.none,
                    "stage": StudentStage.initial_contact,
                    "case_no": str(uuid.uuid4()),
                }
            )

        imported_count = len(student_rows)
        if student_rows:
            await db.execute(insert(Student), student_rows)

        db.add(
            make_operation_log(
                current_user,
                target_student_id=None,
                case_no="",
                action="Excel导入",
                content=f"导入 {imported_count} 条，跳过 {len(skipped_rows)} 条",
            )
        )
        await db.commit()
        return Response.ok(
            {
                "imported": imported_count,
                "success": imported_count,
                "skipped": len(skipped_rows),
                "no_phone": len(no_phone_rows),
                "no_phone_rows": no_phone_rows,
                "skipped_rows": skipped_rows,
                "errors": skipped_rows,
            }
        )
    except (InvalidFileException, BadZipFile, OSError) as exc:
        await db.rollback()
        logger.warning("Excel import parse failed: %s", exc)
        return Response.error(code=1, msg=f"文件解析失败: {exc}")
    except SQLAlchemyError:
        await db.rollback()
        logger.exception("Excel import DB write failed")
        return Response.error(code=1, msg="数据库写入失败")
    except Exception:
        await db.rollback()
        logger.exception("Excel import failed (unexpected)")
        return Response.error(code=1, msg="导入失败，请检查文件格式")
    finally:
        if workbook is not None:
            workbook.close()


@router.get("/template/download")
async def download_import_template():
    """Download Excel import template"""
    wb = Workbook()
    ws = wb.active
    ws.title = "导入模板"

    headers = [
        "姓名",
        "成绩",
        "监护人姓名",
        "监护人电话",
        "监护人2姓名",
        "监护人2电话",
        "学校名称",
        "地域",
    ]
    ws.append(headers)

    # Example data
    ws.append(
        [
            "张三",
            "580",
            "张先生",
            "13900139000",
            "李女士",
            "13700137000",
            "第一中学",
            "福州",
        ]
    )

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 16

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=import_template.xlsx"},
    )


@router.post("")
async def create_student(
    body: StudentCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _require_admin_leads_manage(current_user)
    _require_admin_operation(current_user, ADMIN_OP_STUDENT_CREATE)
    raw = body.model_dump(exclude_unset=True)
    if not is_admin(current_user):
        agent_create_fields = {
            "name",
            "region",
            "score",
            "guardian_name",
            "guardian_phone",
            "guardian2_name",
            "guardian2_phone",
            "school_name",
            "school_address",
            "join_reasons",
        }
        forbidden = sorted(set(raw) - agent_create_fields)
        if forbidden:
            raise HTTPException(status_code=403, detail=f"无权设置字段: {', '.join(forbidden)}")

    status = StudentStatus.not_contacted
    status_detail = ""
    intent_level = IntentLevel.none
    stage = StudentStage.initial_contact
    if is_admin(current_user):
        if body.status:
            try:
                status, implicit_detail = normalize_status_for_write(body.status)
                status_detail = status_detail_for_write(status, implicit_detail)
            except ValueError as e:
                return Response.error(code=1, msg=str(e))
        if body.intent_level:
            try:
                intent_level = _enum_or_error(IntentLevel, body.intent_level, "意向等级")
            except ValueError as e:
                return Response.error(code=1, msg=str(e))
        if body.stage:
            try:
                stage = _enum_or_error(StudentStage, body.stage, "阶段")
            except ValueError as e:
                return Response.error(code=1, msg=str(e))

    assigned_to = body.assigned_to if is_admin(current_user) else current_user.id
    if assigned_to:
        agent_result = await db.execute(
            select(User.id).where(User.id == assigned_to, User.is_active)
        )
        if not agent_result.scalar_one_or_none():
            return Response.error(code=1, msg="话务员不存在或已禁用")

    region_value = body.region
    if not region_value and body.school_name:
        region_value = extract_region(body.school_name)

    guardian_phone, guardian2_phone = _dedupe_contact_phones(
        body.guardian_phone,
        body.guardian2_phone,
    )
    if not _has_any_phone(guardian_phone, guardian2_phone):
        return Response.error(code=1, msg="至少需要一个可拨电话")

    student = Student(
        name=body.name,
        region=region_value,
        assigned_to=assigned_to,
        assigned_at=utcnow() if assigned_to else None,
        status=status,
        status_detail=status_detail,
        intent_level=intent_level,
        stage=stage,
        join_reasons=body.join_reasons or "",
        enrolled_at=body.enrolled_at,
        program=body.program or "",
        deposit=body.deposit,
        score=body.score,
        guardian_name=body.guardian_name or "",
        guardian_phone=guardian_phone,
        guardian2_name=body.guardian2_name or "",
        guardian2_phone=guardian2_phone,
        school_name=body.school_name or "",
        school_address=body.school_address or "",
        need_help=body.need_help or False,
        case_no=str(uuid.uuid4()),
    )
    if student.stage == StudentStage.enrolled:
        student.status = StudentStatus.enrolled
        student.status_detail = ""
        if not student.enrolled_at:
            student.enrolled_at = date.today()

    db.add(student)
    await db.commit()
    await db.refresh(student)
    if student.intent_level == IntentLevel.A:
        asyncio.create_task(
            notify_a_level_change_background(student.id, current_user.name, "create")
        )
    return Response.ok(_student_payload(student))


@router.get("")
async def list_students(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    q: str = Query(""),
    status: str = Query(""),
    status_detail: str = Query(""),
    intent_level: str = Query(""),
    assigned_to: int = Query(None),
    assignment: str = Query(""),
    region: str = Query(""),
    stage: str = Query(""),
    need_help: str = Query(""),
    school_name: str = Query(""),
    active: str = Query(""),
    today_a: str = Query(""),
    missing_phone: str = Query(""),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _require_admin_leads_manage(current_user)
    query = apply_student_scope(select(Student), current_user)
    if q:
        q = q.strip()
        if is_phone_query(q):
            phone_q = normalize_phone(q)
            query = query.where(
                or_(
                    Student.guardian_phone == phone_q,
                    Student.guardian2_phone == phone_q,
                )
            )
        else:
            query = query.where(
                or_(
                    Student.name.contains(q),
                    Student.region.contains(q),
                    Student.school_name.contains(q),
                    Student.guardian_name.contains(q),
                )
            )
    if status:
        # SAEnum 列存的是 enum.name（英文），前端传的是 value（中文）。
        # 转成 enum 实例后 SQLAlchemy 才会正确映射为 name 进 SQL。
        try:
            status_enum = canonical_student_status(StudentStatus(status))
        except ValueError:
            return Response.ok({"total": 0, "page": page, "page_size": page_size, "list": []})
        query = query.where(Student.status.in_(statuses_for_canonical(status_enum)))
    else:
        query = query.where(Student.status.not_in(statuses_for_canonical(StudentStatus.invalid)))
    if status_detail:
        query = query.where(Student.status_detail == status_detail)
    if intent_level:
        try:
            intent_enum = IntentLevel(intent_level)
        except ValueError:
            return Response.ok({"total": 0, "page": page, "page_size": page_size, "list": []})
        query = query.where(Student.intent_level == intent_enum)
    if assignment == "unassigned" or assigned_to == 0:
        if not is_admin(current_user):
            raise HTTPException(status_code=403, detail="无权查看未分配学生")
        query = query.where(Student.assigned_to.is_(None))
    elif assigned_to is not None:
        if not is_admin(current_user) and assigned_to != current_user.id:
            raise HTTPException(status_code=403, detail="无权查看其他坐席的学生")
        query = query.where(Student.assigned_to == assigned_to)
    if region:
        query = query.where(Student.region == region)
    if school_name:
        query = query.where(Student.school_name == school_name)
    if stage:
        try:
            stage_enum = StudentStage(stage)
        except ValueError:
            return Response.ok({"total": 0, "page": page, "page_size": page_size, "list": []})
        query = query.where(Student.stage.in_(_stage_filter_values(stage_enum)))
    if need_help == "1":
        query = query.where(Student.need_help)
    if active == "1":
        query = query.where(Student.status.in_(ACTIVE_TASK_STATUSES))
    if missing_phone == "1":
        query = query.where(
            or_(
                Student.guardian_phone == "",
                Student.guardian_phone.is_(None),
            ),
            or_(
                Student.guardian2_phone == "",
                Student.guardian2_phone.is_(None),
            ),
        )
    if today_a == "1":
        today = today_cst_as_utc()
        tomorrow = today + timedelta(days=1)
        today_a_student_ids = (
            select(OperationLog.target_student_id)
            .where(
                OperationLog.action.in_(["AI分析", "手动评级"]),
                OperationLog.new_status == "A",
                OperationLog.old_status != "A",
                OperationLog.created_at >= today,
                OperationLog.created_at < tomorrow,
                OperationLog.target_student_id.is_not(None),
            )
            .distinct()
        )
        query = query.where(Student.id.in_(today_a_student_ids))

    count_q = select(func.count()).select_from(query.subquery())
    total = (await db.execute(count_q)).scalar()

    query = (
        query.offset((page - 1) * page_size).limit(page_size).order_by(Student.created_at.desc())
    )
    result = await db.execute(query)
    students = result.scalars().all()

    return Response.ok(
        {
            "total": total,
            "page": page,
            "page_size": page_size,
            "list": [_student_payload(s) for s in students],
        }
    )


@router.get("/enrolled")
async def enrolled_students(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """已报名学生列表（含报名信息）"""
    query = (
        select(Student)
        .where(Student.status == StudentStatus.enrolled)
        .order_by(Student.enrolled_at.desc().nullslast())
    )
    query = apply_student_scope(query, current_user)
    count_q = select(func.count()).select_from(query.subquery())
    total = (await db.execute(count_q)).scalar()
    query = query.offset((page - 1) * page_size).limit(page_size)
    result = await db.execute(query)
    students = result.scalars().all()

    # Batch load agent names
    agent_ids = list({s.assigned_to for s in students if s.assigned_to})
    agent_map = {}
    if agent_ids:
        agent_r = await db.execute(select(User.id, User.name).where(User.id.in_(agent_ids)))
        agent_map = dict(agent_r.all())

    data = [
        {
            "id": s.id,
            "name": s.name,
            "region": s.region,
            "program": s.program,
            "deposit": s.deposit,
            "enrolled_at": str(s.enrolled_at) if s.enrolled_at else None,
            "agent_name": agent_map.get(s.assigned_to, ""),
        }
        for s in students
    ]

    deposit_query = apply_student_scope(
        select(func.sum(Student.deposit)).where(Student.status == StudentStatus.enrolled),
        current_user,
    )
    deposit_total = await db.execute(deposit_query)
    total_deposit = deposit_total.scalar() or 0

    return Response.ok(
        {
            "total": total,
            "total_deposit": total_deposit,
            "page": page,
            "page_size": page_size,
            "list": data,
        }
    )


@router.get("/dispatch-regions")
async def list_dispatch_regions(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_page_permission(ADMIN_PAGE_LEADS_MANAGE)),
):
    """获取「未分配且有学校名」学生的区县列表及其未分配人数。"""
    result = await db.execute(
        select(Student.region, func.count(Student.id))
        .where(
            Student.school_name != "",
            Student.region != "",
            Student.assigned_to.is_(None),
        )
        .group_by(Student.region)
        .order_by(func.count(Student.id).desc())
    )
    regions = [{"name": row[0], "count": row[1]} for row in result.all()]
    return Response.ok(regions)


@router.get("/schools")
async def list_schools(
    regions: list[str] = Query(default=[]),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_page_permission(ADMIN_PAGE_LEADS_MANAGE)),
):
    """获取有未分配学生的学校列表及其未分配数量。

    可选 regions：仅统计属于这些区县的未分配学生。
    """
    cleaned_regions = [r.strip() for r in regions if r and r.strip()]
    conditions = [Student.school_name != "", Student.assigned_to.is_(None)]
    if cleaned_regions:
        conditions.append(Student.region.in_(cleaned_regions))
    result = await db.execute(
        select(Student.school_name, func.count(Student.id))
        .where(*conditions)
        .group_by(Student.school_name)
        .order_by(func.count(Student.id).desc())
    )
    schools = [{"name": row[0], "count": row[1]} for row in result.all()]
    return Response.ok(schools)


_CST = timezone(timedelta(hours=8))
DIAL_LOG_DEDUP_SECONDS = 3


async def _get_system_config(db: AsyncSession, key: str, default: str = "") -> str:
    result = await db.execute(select(SystemConfig.value).where(SystemConfig.key == key))
    value = result.scalar_one_or_none()
    return (value or "").strip() or default


def _parse_hhmm(value: str) -> tuple[int, int]:
    try:
        hh, mm = value.split(":")
        return int(hh), int(mm)
    except (ValueError, AttributeError):
        return 0, 0


def _minutes_since_midnight(value: str) -> int:
    hh, mm = _parse_hhmm(value)
    return max(0, min(23, hh)) * 60 + max(0, min(59, mm))


def _is_within_dial_window(current_minutes: int, window_start: str, window_end: str) -> bool:
    start_minutes = _minutes_since_midnight(window_start)
    end_minutes = _minutes_since_midnight(window_end)
    if start_minutes <= end_minutes:
        return start_minutes <= current_minutes <= end_minutes
    return current_minutes >= start_minutes or current_minutes <= end_minutes


@router.get("/phone/{student_id}")
async def get_student_phone(
    student_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    student = await get_accessible_student(db, student_id, current_user)
    _require_admin_operation(current_user, ADMIN_OP_STUDENT_PHONE)

    # 1. 拨号窗口校验
    window_start = await _get_system_config(db, "dial_window_start", "08:00")
    window_end = await _get_system_config(db, "dial_window_end", "21:00")
    max_per_24h_str = await _get_system_config(db, "dial_max_per_24h", "3")
    try:
        max_per_24h = int(max_per_24h_str)
    except ValueError:
        max_per_24h = 3

    now_cst = datetime.now(_CST)
    cur_minutes = now_cst.hour * 60 + now_cst.minute
    if not _is_within_dial_window(cur_minutes, window_start, window_end):
        raise HTTPException(
            status_code=403,
            detail=f"当前为禁拨时段（拨号窗口 {window_start}-{window_end}）",
        )

    # 2. 短时间重复点击同一个拨号按钮时复用本次取号，不重复写 DialLog。
    duplicate_since = utcnow() - timedelta(seconds=DIAL_LOG_DEDUP_SECONDS)
    duplicate_r = await db.execute(
        select(DialLog)
        .where(
            DialLog.student_id == student.id,
            DialLog.agent_id == current_user.id,
            DialLog.dialed_at >= duplicate_since,
        )
        .order_by(DialLog.dialed_at.desc(), DialLog.id.desc())
        .limit(1)
    )
    recent_duplicate = duplicate_r.scalar_one_or_none()
    if recent_duplicate is not None:
        return Response.ok(
            {
                "guardian_phone": student.guardian_phone,
                "guardian2_phone": student.guardian2_phone,
            }
        )

    # 3. 24h 防撞号校验（全局，任何坐席）
    since = utcnow() - timedelta(hours=24)
    count_r = await db.execute(
        select(func.count(DialLog.id)).where(
            DialLog.student_id == student.id,
            DialLog.dialed_at >= since,
        )
    )
    count_24h = count_r.scalar() or 0
    if count_24h >= max_per_24h:
        raise HTTPException(
            status_code=403,
            detail=f"该学生 24h 内已被拨打 {count_24h} 次，达到上限 {max_per_24h}",
        )

    # 4. 通过校验，写 DialLog 并记录操作日志
    db.add(DialLog(student_id=student.id, agent_id=current_user.id))
    db.add(
        make_operation_log(
            current_user,
            student.id,
            student.case_no or "",
            "查看电话",
            content="查看明文电话号码",
        )
    )
    await db.commit()
    return Response.ok(
        {
            "guardian_phone": student.guardian_phone,
            "guardian2_phone": student.guardian2_phone,
        }
    )


@router.put("/dial-duration")
async def update_dial_duration(
    student_id: int = Query(...),
    duration_seconds: int = Query(..., ge=0, le=24 * 60 * 60),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    await get_accessible_student(db, student_id, current_user)

    result = await db.execute(
        select(DialLog)
        .where(DialLog.student_id == student_id, DialLog.agent_id == current_user.id)
        .order_by(DialLog.dialed_at.desc(), DialLog.id.desc())
        .limit(1)
    )
    dial_log = result.scalar_one_or_none()
    if dial_log is None:
        return Response.error(code=1, msg="未找到本次拨号记录")

    dial_log.duration_seconds = duration_seconds
    await db.commit()
    return Response.ok(
        {
            "id": dial_log.id,
            "student_id": dial_log.student_id,
            "duration_seconds": dial_log.duration_seconds,
        }
    )


@router.get("/{student_id}/phone-plain")
async def reveal_student_phone_plain(
    student_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_page_permission(ADMIN_PAGE_LEADS_MANAGE)),
):
    _require_admin_operation(current_user, ADMIN_OP_STUDENT_PHONE)
    student = await get_student_or_404(db, student_id)
    db.add(
        make_operation_log(
            current_user,
            student.id,
            student.case_no or "",
            "查看明文电话",
            content="管理员查看明文电话号码",
        )
    )
    await db.commit()
    return Response.ok(
        {
            "guardian_phone": student.guardian_phone,
            "guardian2_phone": student.guardian2_phone,
        }
    )


@router.get("/{student_id}/intent-timeline")
async def get_intent_timeline(
    student_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    student = await get_accessible_student(db, student_id, current_user)

    result = await db.execute(
        select(Call.id, Call.ai_intent, Call.ai_confidence, Call.agent_id, Call.created_at)
        .where(Call.student_id == student.id, Call.ai_intent != "", Call.ai_intent != "无")
        .order_by(Call.created_at.asc())
    )
    timeline = [
        {
            "call_id": call_id,
            "intent": ai_intent,
            "confidence": ai_confidence,
            "agent_id": agent_id,
            "at": str(created_at),
        }
        for call_id, ai_intent, ai_confidence, agent_id, created_at in result.all()
    ]

    return Response.ok(
        {
            "student_id": student.id,
            "current_intent": str(student.intent_level),
            "created_at": str(student.created_at),
            "timeline": timeline,
        }
    )


class EnrollmentSubStageBody(BaseModel):
    enrollment_substage: str | None = None


@router.put("/{student_id}/enrollment-substage")
async def update_enrollment_substage(
    student_id: int,
    body: EnrollmentSubStageBody,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if not is_admin(current_user):
        raise HTTPException(status_code=403, detail="无权修改报名后状态")
    if not user_has_page_permission(current_user, ADMIN_PAGE_ENROLLMENT_SETTLEMENT):
        raise HTTPException(status_code=403, detail="无权访问该管理模块")
    student = await get_student_or_404(db, student_id)

    old_value = str(student.enrollment_substage) if student.enrollment_substage else ""
    if body.enrollment_substage is None or body.enrollment_substage == "":
        student.enrollment_substage = None
        new_value = ""
    else:
        try:
            student.enrollment_substage = EnrollmentSubStage(body.enrollment_substage)
        except ValueError:
            valid = [e.value for e in EnrollmentSubStage]
            raise HTTPException(
                status_code=400,
                detail=f"无效的报名后状态，合法值：{valid}",
            )
        new_value = body.enrollment_substage

    db.add(
        make_operation_log(
            current_user,
            student.id,
            student.case_no or "",
            "修改报名后状态",
            content=f"{old_value} → {new_value}",
            old_status=old_value,
            new_status=new_value,
        )
    )
    await db.commit()
    await db.refresh(student)
    return Response.ok(
        {
            "enrollment_substage": str(student.enrollment_substage)
            if student.enrollment_substage
            else None,
        }
    )


@router.get("/{student_id}/detail")
async def get_student_detail(
    student_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """聚合学生所有维度信息：基本资料 + 通话 + 备注 + 回访 + 到访 + 意向轨迹。"""
    _require_admin_leads_manage(current_user)
    student = await get_accessible_student(db, student_id, current_user)

    log = LeadViewLog(student_id=student.id, viewer_id=current_user.id)
    db.add(log)

    # 学生基本信息
    payload = _student_payload(student)
    payload["enrollment_substage"] = (
        str(student.enrollment_substage) if student.enrollment_substage else None
    )

    # 通话（最近 50）
    calls_r = await db.execute(
        select(Call, User.name)
        .outerjoin(User, Call.agent_id == User.id)
        .where(Call.student_id == student.id)
        .order_by(Call.created_at.desc())
        .limit(50)
    )
    calls = [
        {
            "id": c.id,
            "agent_id": c.agent_id,
            "agent_name": agent_name or "",
            "duration_seconds": c.duration_seconds,
            "ai_intent": c.ai_intent,
            "ai_confidence": c.ai_confidence,
            "ai_summary": c.ai_summary,
            "ai_reasons": c.ai_reasons,
            "created_at": str(c.created_at),
        }
        for c, agent_name in calls_r.all()
    ]

    # 备注（最近 50）
    notes_r = await db.execute(
        select(Note, User.name)
        .outerjoin(User, Note.agent_id == User.id)
        .where(Note.student_id == student.id)
        .order_by(Note.created_at.desc())
        .limit(50)
    )
    notes = [
        {
            "id": n.id,
            "content": n.content,
            "source": n.source,
            "agent_id": n.agent_id,
            "agent_name": agent_name or "",
            "created_at": str(n.created_at),
            "updated_at": str(n.updated_at),
        }
        for n, agent_name in notes_r.all()
    ]

    # 回访（全部）
    fu_r = await db.execute(
        select(FollowUp, User.name)
        .outerjoin(User, FollowUp.agent_id == User.id)
        .where(FollowUp.student_id == student.id)
        .order_by(FollowUp.follow_up_date.desc())
    )
    follow_ups = [
        {
            "id": f.id,
            "agent_id": f.agent_id,
            "agent_name": agent_name or "",
            "follow_up_date": str(f.follow_up_date),
            "follow_up_type": f.follow_up_type or "",
            "notes": f.notes or "",
            "is_completed": f.is_completed,
            "is_notified": f.is_notified,
            "created_at": str(f.created_at),
        }
        for f, agent_name in fu_r.all()
    ]

    # 到访（全部）
    visits_r = await db.execute(
        select(Visit, User.name)
        .outerjoin(User, Visit.agent_id == User.id)
        .where(Visit.student_id == student.id)
        .order_by(Visit.scheduled_date.desc())
    )
    visits = [
        {
            "id": v.id,
            "agent_id": v.agent_id,
            "agent_name": agent_name or "",
            "visit_type": str(v.visit_type),
            "scheduled_date": str(v.scheduled_date),
            "status": str(v.status),
            "notes": v.notes or "",
            "created_at": str(v.created_at),
        }
        for v, agent_name in visits_r.all()
    ]

    home_visits_r = await db.execute(
        select(HomeVisitTask, User.name)
        .outerjoin(User, HomeVisitTask.creator_agent_id == User.id)
        .where(HomeVisitTask.student_id == student.id)
        .order_by(HomeVisitTask.created_at.desc())
    )
    home_visit_events = [
        {
            "type": "home_visit",
            "id": task.id,
            "title": "申请家访",
            "status": task.status.value,
            "result": task.result.value if task.result else "",
            "operator_name": agent_name or "",
            "occurred_at": str(task.created_at),
            "scheduled_at": str(task.scheduled_at) if task.scheduled_at else None,
            "summary": task.address or task.notes or "",
        }
        for task, agent_name in home_visits_r.all()
    ]

    campus_visits_r = await db.execute(
        select(CampusVisitTask, User.name)
        .outerjoin(User, CampusVisitTask.creator_user_id == User.id)
        .where(CampusVisitTask.student_id == student.id)
        .order_by(CampusVisitTask.created_at.desc())
    )
    campus_visit_events = [
        {
            "type": "campus_visit",
            "id": task.id,
            "title": "预约到校",
            "status": task.status.value,
            "result": task.result.value if task.result else "",
            "operator_name": user_name or "",
            "occurred_at": str(task.created_at),
            "scheduled_at": str(task.appointment_at) if task.appointment_at else None,
            "summary": task.current_concerns or task.notes or "",
        }
        for task, user_name in campus_visits_r.all()
    ]

    enrollments_r = await db.execute(
        select(EnrollmentRecord, User.name)
        .outerjoin(User, EnrollmentRecord.attributed_agent_id == User.id)
        .where(EnrollmentRecord.student_id == student.id)
        .order_by(EnrollmentRecord.enrolled_at.desc())
    )
    enrollment_events = [
        {
            "type": "enrollment",
            "id": record.id,
            "title": "报名登记",
            "status": record.settlement_status.value,
            "result": record.source.value,
            "operator_name": agent_name or "",
            "occurred_at": str(record.enrolled_at),
            "scheduled_at": None,
            "summary": record.enrolled_program or record.intent_program or "",
        }
        for record, agent_name in enrollments_r.all()
    ]

    admissions_timeline = sorted(
        home_visit_events + campus_visit_events + enrollment_events,
        key=lambda item: item.get("occurred_at") or "",
        reverse=True,
    )

    # 意向轨迹：合并 AI 分析（Call）和手动评级（OperationLog）
    intent_r = await db.execute(
        select(Call.id, Call.ai_intent, Call.ai_confidence, Call.agent_id, Call.created_at)
        .where(Call.student_id == student.id, Call.ai_intent != "", Call.ai_intent != "无")
        .order_by(Call.created_at.asc())
    )
    ai_events = [
        {
            "source": "ai",
            "intent_level": ai_intent,
            "confidence": ai_conf,
            "agent_id": aid,
            "created_at": str(created_at),
        }
        for cid, ai_intent, ai_conf, aid, created_at in intent_r.all()
    ]

    manual_r = await db.execute(
        select(OperationLog)
        .where(
            OperationLog.target_student_id == student.id,
            OperationLog.action == "手动评级",
        )
        .order_by(OperationLog.created_at.asc())
    )
    manual_events = [
        {
            "source": "manual",
            "intent_level": log.new_status or "无",
            "old_intent": log.old_status or "",
            "operator_name": log.operator_name or "",
            "created_at": str(log.created_at),
        }
        for log in manual_r.scalars().all()
    ]

    # 合并并按时间排序
    intent_timeline = sorted(
        ai_events + manual_events,
        key=lambda x: x.get("created_at", ""),
    )

    await db.commit()

    return Response.ok(
        {
            "student": payload,
            "calls": calls,
            "notes": notes,
            "follow_ups": follow_ups,
            "visits": visits,
            "admissions_timeline": admissions_timeline,
            "intent_timeline": intent_timeline,
        }
    )


@router.get("/{student_id}")
async def get_student(
    student_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _require_admin_leads_manage(current_user)
    student = await get_accessible_student(db, student_id, current_user)

    log = LeadViewLog(student_id=student.id, viewer_id=current_user.id)
    db.add(log)
    await db.commit()

    return Response.ok(_student_payload(student))


@router.put("/{student_id}")
async def update_student(
    student_id: int,
    body: StudentUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _require_admin_leads_manage(current_user)
    _require_admin_operation(current_user, ADMIN_OP_STUDENT_EDIT)
    student = await get_accessible_student(db, student_id, current_user)
    raw = body.model_dump(exclude_unset=True)
    # invalid_reason is persisted as status_detail for 无效 so admins can filter by reason.
    invalid_reason = (raw.pop("invalid_reason", None) or "").strip()
    if not raw:
        return Response.ok(_student_payload(student))

    allowed_fields = (
        ADMIN_STUDENT_UPDATE_FIELDS if is_admin(current_user) else AGENT_STUDENT_UPDATE_FIELDS
    )
    forbidden = sorted(set(raw) - allowed_fields)
    if forbidden:
        raise HTTPException(status_code=403, detail=f"无权修改字段: {', '.join(forbidden)}")

    old_intent = student.intent_level
    old_status = student.status
    old_status_detail = student.status_detail or ""
    old_stage = student.stage
    old_assigned = student.assigned_to
    next_guardian_phone = student.guardian_phone
    next_guardian2_phone = student.guardian2_phone
    was_enrolled = _is_enrolled_student(student)
    for k, v in raw.items():
        if k == "status" and v is not None:
            try:
                v, implicit_status_detail = normalize_status_for_write(v)
                if was_enrolled and canonical_student_status(v) != StudentStatus.enrolled:
                    return Response.error(code=1, msg="已报名学生不能通过普通编辑改回非报名状态")
                student.status_detail = status_detail_for_write(
                    v,
                    implicit_status_detail,
                    invalid_reason,
                )
            except ValueError as e:
                return Response.error(code=1, msg=str(e))
        elif k == "stage" and v is not None:
            try:
                v = _enum_or_error(StudentStage, v, "阶段")
                if was_enrolled and v != StudentStage.enrolled:
                    return Response.error(code=1, msg="已报名学生不能通过普通编辑改回非报名状态")
            except ValueError as e:
                return Response.error(code=1, msg=str(e))
        elif k == "intent_level" and v is not None:
            try:
                v = _enum_or_error(IntentLevel, v, "意向等级")
            except ValueError as e:
                return Response.error(code=1, msg=str(e))
        elif k in {"guardian_phone", "guardian2_phone"} and v is not None:
            v = normalize_phone(v)
            if k == "guardian_phone":
                next_guardian_phone = v
            else:
                next_guardian2_phone = v
        setattr(student, k, v)

    if was_enrolled and not _is_enrolled_student(student):
        return Response.error(code=1, msg="已报名学生不能通过普通编辑改回非报名状态")

    if {"guardian_phone", "guardian2_phone"} & set(raw):
        next_guardian_phone, next_guardian2_phone = _dedupe_contact_phones(
            next_guardian_phone,
            next_guardian2_phone,
        )
        if not _has_any_phone(next_guardian_phone, next_guardian2_phone):
            return Response.error(code=1, msg="至少需要一个可拨电话")
        student.guardian_phone = next_guardian_phone
        student.guardian2_phone = next_guardian2_phone

    if student.stage == StudentStage.enrolled:
        student.status = StudentStatus.enrolled
        student.status_detail = ""
        if not student.enrolled_at:
            student.enrolled_at = date.today()
    elif student.status == StudentStatus.enrolled:
        student.stage = StudentStage.enrolled
        student.status_detail = ""
        if not student.enrolled_at:
            student.enrolled_at = date.today()

    intent_changed = "intent_level" in raw and old_intent != student.intent_level
    status_changed = old_status != student.status
    status_detail_changed = old_status_detail != (student.status_detail or "")
    stage_changed = old_stage != student.stage
    assigned_changed = "assigned_to" in raw and old_assigned != student.assigned_to

    if status_changed or status_detail_changed:
        if _is_call_result_write(
            student.status, student.status_detail
        ) and not _allows_call_result_backfill_without_recent_dial(old_status):
            await require_recent_agent_dial(db, student.id, current_user)

    if intent_changed:
        db.add(
            make_operation_log(
                current_user,
                student.id,
                student.case_no or "",
                "手动评级",
                content=f"意向 {old_intent} → {student.intent_level}",
                old_status=str(old_intent),
                new_status=str(student.intent_level),
                note_content="",
            )
        )

    if status_changed or status_detail_changed or stage_changed or assigned_changed:
        parts = []
        if status_changed:
            parts.append(
                f"状态 {canonical_status_value(old_status)} → "
                f"{canonical_status_value(student.status)}"
            )
        if status_detail_changed and student.status_detail:
            parts.append(f"结果/原因：{student.status_detail}")
        if student.status == StudentStatus.invalid and student.status_detail:
            parts.append(f"无效原因：{student.status_detail}")
        if stage_changed:
            parts.append(f"阶段 {old_stage} → {student.stage}")
        if assigned_changed:
            parts.append(f"分配 {old_assigned} → {student.assigned_to}")
        db.add(
            make_operation_log(
                current_user,
                student.id,
                student.case_no or "",
                "修改状态" if status_changed else "修改信息",
                content="; ".join(parts),
                old_status=canonical_status_value(old_status) if status_changed else "",
                new_status=canonical_status_value(student.status) if status_changed else "",
                note_content=student.status_detail
                if (student.status == StudentStatus.invalid and student.status_detail)
                else "",
            )
        )

    await db.commit()
    await db.refresh(student)
    if intent_changed and student.intent_level == IntentLevel.A:
        asyncio.create_task(
            notify_a_level_change_background(student.id, current_user.name, "manual")
        )
    return Response.ok(_student_payload(student))


@router.put("/{student_id}/stage")
async def update_stage(
    student_id: int,
    body: StageUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _require_admin_leads_manage(current_user)
    student = await get_accessible_student(db, student_id, current_user)
    old_stage = student.stage
    old_status = student.status

    try:
        new_stage = _enum_or_error(StudentStage, body.stage, "阶段")
    except ValueError as e:
        return Response.error(msg=str(e))
    if _is_enrolled_student(student) and new_stage != StudentStage.enrolled:
        return Response.error(code=1, msg="已报名学生不能通过普通编辑改回非报名状态")
    student.stage = new_stage

    # Auto-update status when stage is "已报名"
    if new_stage == StudentStage.enrolled:
        student.status = StudentStatus.enrolled
        student.status_detail = ""
        if not student.enrolled_at:
            student.enrolled_at = date.today()

    stage_changed = old_stage != student.stage
    status_changed = old_status != student.status
    if stage_changed or status_changed:
        parts = []
        if stage_changed:
            parts.append(f"阶段 {old_stage} → {student.stage}")
        if status_changed:
            parts.append(
                f"状态 {canonical_status_value(old_status)} → "
                f"{canonical_status_value(student.status)}"
            )
        db.add(
            make_operation_log(
                current_user,
                student.id,
                student.case_no or "",
                "修改状态" if status_changed else "修改信息",
                content="; ".join(parts),
                old_status=canonical_status_value(old_status) if status_changed else "",
                new_status=canonical_status_value(student.status) if status_changed else "",
            )
        )

    await db.commit()
    await db.refresh(student)
    return Response.ok(
        {"stage": _display_stage(student.stage), "status": canonical_status_value(student.status)}
    )


@router.put("/{student_id}/enroll")
async def set_enroll_info(
    student_id: int,
    body: EnrollInfo,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if not is_admin(current_user):
        raise HTTPException(status_code=403, detail="无权设置报名信息")
    if not user_has_page_permission(current_user, ADMIN_PAGE_ENROLLMENT_SETTLEMENT):
        raise HTTPException(status_code=403, detail="无权访问该管理模块")
    _require_admin_operation(current_user, ADMIN_OP_ENROLLMENT_SETTLEMENT)
    student = await get_student_or_404(db, student_id)
    old_status = student.status
    old_stage = student.stage
    old_enrolled_at = student.enrolled_at
    old_program = student.program
    old_deposit = student.deposit

    student.enrolled_at = body.enrolled_at or date.today()
    student.program = body.program
    student.deposit = body.deposit
    student.status = StudentStatus.enrolled
    student.status_detail = ""
    student.stage = StudentStage.enrolled
    if student.enrollment_substage is None:
        student.enrollment_substage = EnrollmentSubStage.deposit_pending

    parts = [
        f"状态 {canonical_status_value(old_status)} → {canonical_status_value(student.status)}",
        f"阶段 {old_stage} → {student.stage}",
        f"报名日 {old_enrolled_at or '-'} → {student.enrolled_at}",
    ]
    if old_program != student.program:
        parts.append(f"专业 {old_program or '-'} → {student.program or '-'}")
    if old_deposit != student.deposit:
        old_deposit_text = old_deposit if old_deposit is not None else "-"
        new_deposit_text = student.deposit if student.deposit is not None else "-"
        parts.append(f"定金 {old_deposit_text} → {new_deposit_text}")
    db.add(
        make_operation_log(
            current_user,
            student.id,
            student.case_no or "",
            "报名登记",
            content="; ".join(parts),
            old_status=canonical_status_value(old_status),
            new_status=canonical_status_value(student.status),
        )
    )

    await db.commit()
    await db.refresh(student)
    return Response.ok(
        {
            "enrolled_at": str(student.enrolled_at),
            "program": student.program,
            "deposit": student.deposit,
            "enrollment_substage": str(student.enrollment_substage)
            if student.enrollment_substage
            else None,
        }
    )


@router.put("/{student_id}/extend")
async def extend_expiry(
    student_id: int,
    days: int = Query(15, ge=1, le=90),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    """过期逻辑已暂时停用，保留接口以兼容旧前端/脚本调用。"""
    return Response.ok({"expired_at": None, "disabled": True})


class AssignReq(BaseModel):
    student_ids: list[int]
    agent_id: int


class SchoolAssignReq(BaseModel):
    school_name: str = Field(..., min_length=1)
    agent_ids: list[int] = Field(..., min_length=1)
    regions: list[str] = Field(default_factory=list)

    @field_validator("school_name")
    @classmethod
    def normalize_school_name(cls, value: str) -> str:
        value = value.strip()
        if not value:
            raise ValueError("请选择学校")
        return value

    @field_validator("regions")
    @classmethod
    def normalize_regions(cls, value: list[str]) -> list[str]:
        return [region.strip() for region in value if isinstance(region, str) and region.strip()]


def _add_assignment_logs(
    db: AsyncSession,
    current_user: User,
    students: list[Student],
    assigned_by_student_id: dict[int, int],
    *,
    action: str,
    content_prefix: str = "分配给话务员",
    batch_id: str = "",
    old_assignment_by_student_id: dict[int, tuple[int | None, datetime | None]] | None = None,
    assigned_at_by_student_id: dict[int, datetime] | None = None,
):
    for student in students:
        agent_id = assigned_by_student_id.get(student.id)
        if agent_id is None:
            continue
        old_agent_id, old_assigned_at = (old_assignment_by_student_id or {}).get(
            student.id,
            (None, None),
        )
        new_assigned_at = (assigned_at_by_student_id or {}).get(student.id)
        db.add(
            make_operation_log(
                current_user,
                student.id,
                student.case_no or "",
                action,
                content=f"{content_prefix} {agent_id}",
                old_status=assignment_state_label(old_agent_id),
                new_status=assignment_state_label(agent_id),
                note_content=make_assignment_rollback_note(
                    old_assigned_to=old_agent_id,
                    old_assigned_at=old_assigned_at,
                    new_assigned_to=agent_id,
                    new_assigned_at=new_assigned_at,
                ),
                batch_id=batch_id,
            )
        )


def _student_names_preview(students: list[Student], limit: int = 5) -> str:
    names = [student.name or f"学生{student.id}" for student in students[:limit]]
    suffix = f" 等 {len(students)} 人" if len(students) > limit else ""
    return "、".join(names) + suffix


def _add_batch_summary_log(
    db: AsyncSession,
    current_user: User,
    *,
    action: str,
    content: str,
    batch_id: str = "",
):
    db.add(
        make_operation_log(
            current_user,
            target_student_id=None,
            case_no="",
            action=action,
            content=content,
            batch_id=batch_id,
        )
    )


@router.post("/assign")
async def assign_students(
    body: AssignReq,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_operation_permission(ADMIN_OP_STUDENT_ASSIGN)),
):
    if not body.student_ids:
        return Response.error(code=1, msg="student_ids不能为空")

    agent_result = await db.execute(select(User).where(User.id == body.agent_id, User.is_active))
    if not agent_result.scalar_one_or_none():
        return Response.error(code=1, msg="话务员不存在或已禁用")

    students_result = await db.execute(select(Student).where(Student.id.in_(body.student_ids)))
    students = students_result.scalars().all()
    if not students:
        return Response.error(code=1, msg="未找到指定的学生")
    enrolled_students = [student for student in students if _is_enrolled_student(student)]
    if enrolled_students:
        names = "、".join(student.name for student in enrolled_students[:3])
        suffix = f" 等 {len(enrolled_students)} 人" if len(enrolled_students) > 3 else ""
        return Response.error(code=1, msg=f"已报名学生不能重新分配：{names}{suffix}")

    now = utcnow()
    batch_id = make_batch_id("assign")
    old_assignment_by_student_id = {
        student.id: (student.assigned_to, student.assigned_at) for student in students
    }
    assigned_at_by_student_id = {student.id: now for student in students}
    await db.execute(
        update(Student)
        .where(Student.id.in_(body.student_ids))
        .values(assigned_to=body.agent_id, assigned_at=now)
    )
    _add_assignment_logs(
        db,
        current_user,
        students,
        {student.id: body.agent_id for student in students},
        action="手动分配",
        batch_id=batch_id,
        old_assignment_by_student_id=old_assignment_by_student_id,
        assigned_at_by_student_id=assigned_at_by_student_id,
    )
    _add_batch_summary_log(
        db,
        current_user,
        action="批量分配",
        content=(
            f"共 {len(students)} 名学生分配给话务员 {body.agent_id}；"
            f"样例：{_student_names_preview(students)}"
        ),
        batch_id=batch_id,
    )
    await db.commit()

    return Response.ok(
        {"assigned_count": len(students), "agent_id": body.agent_id, "batch_id": batch_id}
    )


@router.post("/auto-assign")
async def auto_assign(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_operation_permission(ADMIN_OP_STUDENT_ASSIGN)),
):
    agent_result = await db.execute(select(User).where(User.is_active, User.role == UserRole.agent))
    agents = agent_result.scalars().all()
    if not agents:
        return Response.error(code=1, msg="没有可用的话务员")

    load = {}
    for a in agents:
        cnt = await db.execute(
            select(func.count(Student.id)).where(
                Student.assigned_to == a.id,
                Student.status.not_in(TERMINAL_STUDENT_STATUSES),
            )
        )
        load[a.id] = cnt.scalar() or 0

    unassigned_result = await db.execute(
        select(Student)
        .where(
            Student.assigned_to.is_(None),
            Student.status.not_in(TERMINAL_STUDENT_STATUSES),
        )
        .order_by(Student.created_at.asc())
    )
    unassigned = unassigned_result.scalars().all()
    if not unassigned:
        return Response.ok({"message": "没有未分配的学生", "distribution": {}})

    distribution = {a.id: 0 for a in agents}
    now = utcnow()
    batch_id = make_batch_id("auto-assign")
    old_assignment_by_student_id = {
        student.id: (student.assigned_to, student.assigned_at) for student in unassigned
    }
    assigned_at_by_student_id = {student.id: now for student in unassigned}
    by_agent: dict[int, list[int]] = {}
    assigned_by_student_id: dict[int, int] = {}
    for student in unassigned:
        min_agent_id = min(load, key=load.get)
        by_agent.setdefault(min_agent_id, []).append(student.id)
        assigned_by_student_id[student.id] = min_agent_id
        load[min_agent_id] += 1
        distribution[min_agent_id] += 1

    for agent_id, ids in by_agent.items():
        if not ids:
            continue
        await db.execute(
            update(Student).where(Student.id.in_(ids)).values(assigned_to=agent_id, assigned_at=now)
        )

    _add_assignment_logs(
        db,
        current_user,
        unassigned,
        assigned_by_student_id,
        action="自动分配",
        batch_id=batch_id,
        old_assignment_by_student_id=old_assignment_by_student_id,
        assigned_at_by_student_id=assigned_at_by_student_id,
    )
    distribution_text = ", ".join(
        f"{a.id}:{distribution.get(a.id, 0)}" for a in agents if distribution.get(a.id, 0) > 0
    )
    _add_batch_summary_log(
        db,
        current_user,
        action="自动分配汇总",
        content=(
            f"自动均摊未分配线索，共 {len(unassigned)} 名；"
            f"分布：{distribution_text}；"
            f"样例：{_student_names_preview(unassigned)}"
        ),
        batch_id=batch_id,
    )
    await db.commit()
    # 按 agent_id 聚合返回，避免重名话务员被合并（name 非唯一）
    result = [
        {"agent_id": a.id, "name": a.name, "count": distribution.get(a.id, 0)}
        for a in agents
        if distribution.get(a.id, 0) > 0
    ]

    return Response.ok(
        {"total_assigned": len(unassigned), "distribution": result, "batch_id": batch_id}
    )


@router.post("/region-assign")
async def region_assign(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_operation_permission(ADMIN_OP_STUDENT_ASSIGN)),
):
    agent_result = await db.execute(select(User).where(User.is_active, User.role == UserRole.agent))
    agents = agent_result.scalars().all()
    if not agents:
        return Response.error(code=1, msg="没有可用的话务员")

    # 一个地区可被多个话务员服务；命中地区时按当前活跃负载最小者分配
    region_map: dict[str, list[User]] = {}
    for a in agents:
        if a.service_regions:
            for r in a.service_regions.replace("，", ",").split(","):
                r = r.strip()
                if r:
                    region_map.setdefault(r, []).append(a)

    # 活跃负载基线（排除终态学生），避免历史已报名/已过期影响公平
    load = {}
    for a in agents:
        cnt = await db.execute(
            select(func.count(Student.id)).where(
                Student.assigned_to == a.id,
                Student.status.not_in(TERMINAL_STUDENT_STATUSES),
            )
        )
        load[a.id] = cnt.scalar() or 0
    agent_name_by_id = {a.id: a.name for a in agents}

    unassigned_result = await db.execute(
        select(Student)
        .where(
            Student.assigned_to.is_(None),
            Student.status.not_in(TERMINAL_STUDENT_STATUSES),
        )
        .order_by(Student.created_at.asc())
    )
    unassigned = unassigned_result.scalars().all()

    distribution = {a.name: {"matched": 0, "fallback": 0} for a in agents}
    now = utcnow()
    batch_id = make_batch_id("region-assign")
    old_assignment_by_student_id = {
        student.id: (student.assigned_to, student.assigned_at) for student in unassigned
    }
    assigned_at_by_student_id = {student.id: now for student in unassigned}
    total_assigned = 0
    by_agent: dict[int, list[int]] = {}
    assigned_by_student_id: dict[int, int] = {}

    for student in unassigned:
        matched_candidates = region_map.get(student.region or "", [])
        if matched_candidates:
            # 同地区多话务员：选负载最小者
            chosen = min(matched_candidates, key=lambda a: load[a.id])
            agent_id = chosen.id
            distribution[chosen.name]["matched"] += 1
        else:
            agent_id = min(load, key=load.get)
            name = agent_name_by_id.get(agent_id, "")
            if name:
                distribution[name]["fallback"] += 1

        by_agent.setdefault(agent_id, []).append(student.id)
        assigned_by_student_id[student.id] = agent_id
        load[agent_id] += 1
        total_assigned += 1

    for agent_id, ids in by_agent.items():
        if not ids:
            continue
        await db.execute(
            update(Student).where(Student.id.in_(ids)).values(assigned_to=agent_id, assigned_at=now)
        )

    _add_assignment_logs(
        db,
        current_user,
        unassigned,
        assigned_by_student_id,
        action="区域分配",
        batch_id=batch_id,
        old_assignment_by_student_id=old_assignment_by_student_id,
        assigned_at_by_student_id=assigned_at_by_student_id,
    )
    _add_batch_summary_log(
        db,
        current_user,
        action="区域分配汇总",
        content=(
            f"区域分配未分配线索，共 {total_assigned} 名；"
            f"样例：{_student_names_preview(unassigned)}"
        ),
        batch_id=batch_id,
    )
    await db.commit()
    return Response.ok(
        {
            "total_assigned": total_assigned,
            "distribution": distribution,
            "batch_id": batch_id,
        }
    )


@router.post("/school-assign")
async def school_assign(
    body: SchoolAssignReq,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_operation_permission(ADMIN_OP_STUDENT_ASSIGN)),
):
    """按学校分发：选学校、可选区县过滤、选多个话务员ID、轮询分配"""
    school = body.school_name
    agent_ids = body.agent_ids
    regions = body.regions

    # 验证话务员存在
    agents_result = await db.execute(
        select(User).where(User.id.in_(agent_ids), User.is_active, User.role == UserRole.agent)
    )
    agents = agents_result.scalars().all()
    if not agents:
        return Response.error(code=1, msg="没有可用的话务员")

    # 查出该学校未分配的学生（可选按区县过滤）
    conditions = [
        Student.school_name == school,
        Student.assigned_to.is_(None),
        Student.status.not_in(TERMINAL_STUDENT_STATUSES),
    ]
    if regions:
        conditions.append(Student.region.in_(regions))
    students_result = await db.execute(
        select(Student).where(*conditions).order_by(Student.created_at.asc())
    )
    students = students_result.scalars().all()
    if not students:
        return Response.error(
            code=1,
            msg="该学校在所选区县下没有未分配的学生" if regions else "该学校没有未分配的学生",
        )

    now = utcnow()
    batch_id = make_batch_id("school-assign")
    old_assignment_by_student_id = {
        student.id: (student.assigned_to, student.assigned_at) for student in students
    }
    assigned_at_by_student_id = {student.id: now for student in students}
    by_agent: dict[int, list[int]] = {}
    agent_id_list = [a.id for a in agents]
    counts = {a_id: 0 for a_id in agent_id_list}

    for student in students:
        # 轮询：选当前分配数量最少的话务员
        min_agent_id = min(counts, key=counts.get)
        by_agent.setdefault(min_agent_id, []).append(student.id)
        counts[min_agent_id] += 1

    for agent_id, ids in by_agent.items():
        await db.execute(
            update(Student).where(Student.id.in_(ids)).values(assigned_to=agent_id, assigned_at=now)
        )

    _add_assignment_logs(
        db,
        current_user,
        students,
        {student_id: agent_id for agent_id, ids in by_agent.items() for student_id in ids},
        action="学校分配",
        content_prefix=f"学校「{school}」分配给话务员",
        batch_id=batch_id,
        old_assignment_by_student_id=old_assignment_by_student_id,
        assigned_at_by_student_id=assigned_at_by_student_id,
    )
    _add_batch_summary_log(
        db,
        current_user,
        action="学校分配汇总",
        content=(
            f"学校「{school}」分发，共 {len(students)} 名；"
            f"区县：{('、'.join(regions) if regions else '全部')}；"
            f"话务员：{', '.join(str(a.id) for a in agents)}；"
            f"样例：{_student_names_preview(students)}"
        ),
        batch_id=batch_id,
    )
    await db.commit()
    return Response.ok(
        {
            "total_assigned": len(students),
            "distribution": {f"agent_{a_id}": len(ids) for a_id, ids in by_agent.items()},
            "batch_id": batch_id,
        }
    )


@router.post("/{student_id}/need-help")
async def toggle_need_help(
    student_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    student = await get_accessible_student(db, student_id, current_user)
    student.need_help = not student.need_help
    db.add(
        make_operation_log(
            current_user,
            student.id,
            student.case_no,
            "标记协助" if student.need_help else "取消协助",
            content="需要协助" if student.need_help else "取消协助标记",
        )
    )
    await db.commit()
    return Response.ok({"need_help": student.need_help})


@router.delete("/{student_id}")
async def delete_student(
    student_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_operation_permission(ADMIN_OP_STUDENT_DELETE)),
):
    student = await get_student_or_404(db, student_id)
    db.add(
        make_operation_log(
            current_user,
            student.id,
            student.case_no or "",
            "删除线索",
            content=f"删除学生 {student.name}（含通话/备注/回访/到访/查看日志）",
        )
    )
    for model in (Call, Note, FollowUp, LeadViewLog, Visit, DialLog):
        await db.execute(delete(model).where(model.student_id == student_id))
    await db.delete(student)
    await db.commit()
    return Response.ok(msg="删除成功")


@router.get("/agent/settings")
async def agent_settings(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_agent),
):
    """返回话务员端需要的非敏感系统配置。"""
    from app.routers.admin import get_config_value

    dial_max_str = await get_config_value(db, "dial_max_per_24h", "3")
    try:
        dial_max = max(1, int(dial_max_str))
    except (ValueError, TypeError):
        dial_max = 3
    return Response.ok({"dial_max_per_24h": dial_max})
