import logging
import uuid
from io import BytesIO
from itertools import chain
from zipfile import BadZipFile

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from sqlalchemy import insert, or_, select
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.ext.asyncio import AsyncSession

from app.auth import (
    ADMIN_OP_STUDENT_IMPORT,
    ADMIN_PAGE_LEADS_MANAGE,
    require_page_permission,
    user_has_operation_permission,
)
from app.database import get_db
from app.models import IntentLevel, Student, StudentStage, StudentStatus, User
from app.schemas import Response
from app.student_import import (
    MAX_STUDENT_IMPORT_BYTES,
    build_import_header_map,
    is_empty_import_row,
    parse_import_row,
)
from app.utils import make_operation_log, utcnow

router = APIRouter(prefix="/api/students", tags=["学生"])
logger = logging.getLogger(__name__)


def _require_admin_operation(current_user: User, permission: str) -> None:
    if not user_has_operation_permission(current_user, permission):
        raise HTTPException(status_code=403, detail="无权执行该操作")


@router.post("/import")
async def import_students_excel(
    file: UploadFile = File(...),
    default_agent_id: int | None = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_page_permission(ADMIN_PAGE_LEADS_MANAGE)),
):
    _require_admin_operation(current_user, ADMIN_OP_STUDENT_IMPORT)
    filename = file.filename or ""
    if not filename.lower().endswith(".xlsx"):
        return Response.error(code=1, msg="仅支持 .xlsx 文件")

    workbook = None
    try:
        contents = await file.read()
        if len(contents) > MAX_STUDENT_IMPORT_BYTES:
            max_mb = MAX_STUDENT_IMPORT_BYTES // (1024 * 1024)
            return Response.error(code=1, msg=f"文件过大，请小于 {max_mb}MB")
        if not contents:
            return Response.error(code=1, msg="上传文件为空")

        if default_agent_id is not None:
            agent_result = await db.execute(
                select(User.id).where(User.id == default_agent_id, User.is_active)
            )
            if not agent_result.scalar_one_or_none():
                return Response.error(code=1, msg="默认话务员不存在或已禁用")

        workbook = load_workbook(filename=BytesIO(contents), read_only=True, data_only=True)
        ws = workbook.active
        rows = ws.iter_rows(values_only=True)
        header_row = next(rows, None)
        if not header_row:
            return Response.error(code=1, msg="Excel 文件没有表头")

        header_map = build_import_header_map(header_row)
        data_start_row = 2
        if "name" not in header_map:
            rows = chain([header_row], rows)
            header_map = {}
            data_start_row = 1

        parsed_rows = []  # (row_idx, parsed_dict)
        skipped_rows = []
        assigned_at = utcnow() if default_agent_id is not None else None

        for row_idx, row in enumerate(rows, start=data_start_row):
            if is_empty_import_row(row):
                continue

            parsed, error = parse_import_row(row, header_map)
            if error:
                skipped_rows.append({"row": row_idx, "reason": error})
                continue

            parsed_rows.append((row_idx, parsed))

        # Batch-check for duplicate phone numbers against existing DB records.
        # 同一行的 guardian_phone / guardian2_phone 都纳入查重；库里命中任一字段视为已存在。
        phones_to_check: set[str] = set()
        for _, p in parsed_rows:
            for key in ("guardian_phone", "guardian2_phone"):
                val = p.get(key) or ""
                if val:
                    phones_to_check.add(val)

        existing_phones: set[str] = set()
        if phones_to_check:
            existing_r = await db.execute(
                select(Student.guardian_phone, Student.guardian2_phone).where(
                    or_(
                        Student.guardian_phone.in_(phones_to_check),
                        Student.guardian2_phone.in_(phones_to_check),
                    )
                )
            )
            for g1, g2 in existing_r.all():
                if g1:
                    existing_phones.add(g1)
                if g2:
                    existing_phones.add(g2)

        seen_in_file: set[str] = set()
        student_rows = []
        no_phone_rows = []
        for row_idx, parsed in parsed_rows:
            phone = parsed.get("guardian_phone", "") or ""
            phone2 = parsed.get("guardian2_phone", "") or ""
            row_phones = [p for p in (phone, phone2) if p]

            if not row_phones:
                item = {
                    "row": row_idx,
                    "name": parsed["name"],
                    "reason": "无电话数据",
                }
                no_phone_rows.append(item)
                skipped_rows.append(item)
                continue

            dup_db = next((p for p in row_phones if p in existing_phones), None)
            if dup_db:
                skipped_rows.append(
                    {"row": row_idx, "reason": f"手机号已存在（库中已有该学员）: {dup_db}"}
                )
                continue
            dup_file = next((p for p in row_phones if p in seen_in_file), None)
            if dup_file:
                skipped_rows.append(
                    {"row": row_idx, "reason": f"手机号在本次导入中重复: {dup_file}"}
                )
                continue
            seen_in_file.update(row_phones)

            student_rows.append(
                {
                    "name": parsed["name"],
                    "region": parsed["region"],
                    "assigned_to": default_agent_id,
                    "assigned_at": assigned_at,
                    "score": parsed["score"],
                    "guardian_name": parsed["guardian_name"],
                    "guardian_phone": phone,
                    "guardian2_name": parsed["guardian2_name"],
                    "guardian2_phone": phone2,
                    "school_name": parsed["school_name"],
                    "school_address": parsed["school_address"],
                    "program": parsed["program"],
                    "join_reasons": parsed["join_reasons"],
                    "status": StudentStatus.not_contacted,
                    "status_detail": "",
                    "intent_level": IntentLevel.none,
                    "stage": StudentStage.initial_contact,
                    "case_no": str(uuid.uuid4()),
                }
            )

        imported_count = len(student_rows)
        if student_rows:
            await db.execute(insert(Student), student_rows)

        db.add(
            make_operation_log(
                current_user,
                target_student_id=None,
                case_no="",
                action="Excel导入",
                content=f"导入 {imported_count} 条，跳过 {len(skipped_rows)} 条",
            )
        )
        await db.commit()
        return Response.ok(
            {
                "imported": imported_count,
                "success": imported_count,
                "skipped": len(skipped_rows),
                "no_phone": len(no_phone_rows),
                "no_phone_rows": no_phone_rows,
                "skipped_rows": skipped_rows,
                "errors": skipped_rows,
            }
        )
    except (InvalidFileException, BadZipFile, OSError) as exc:
        await db.rollback()
        logger.warning("Excel import parse failed: %s", exc)
        return Response.error(code=1, msg=f"文件解析失败: {exc}")
    except SQLAlchemyError:
        await db.rollback()
        logger.exception("Excel import DB write failed")
        return Response.error(code=1, msg="数据库写入失败")
    except Exception:
        await db.rollback()
        logger.exception("Excel import failed (unexpected)")
        return Response.error(code=1, msg="导入失败，请检查文件格式")
    finally:
        if workbook is not None:
            workbook.close()


@router.get("/template/download")
async def download_import_template():
    """Download Excel import template"""
    wb = Workbook()
    ws = wb.active
    ws.title = "导入模板"

    headers = [
        "姓名",
        "成绩",
        "监护人姓名",
        "监护人电话",
        "监护人2姓名",
        "监护人2电话",
        "学校名称",
        "地域",
    ]
    ws.append(headers)

    # Example data
    ws.append(
        [
            "张三",
            "580",
            "张先生",
            "13900139000",
            "李女士",
            "13700137000",
            "第一中学",
            "福州",
        ]
    )

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 16

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=import_template.xlsx"},
    )
